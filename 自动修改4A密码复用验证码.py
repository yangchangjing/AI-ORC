import pyautogui
import time
import re
import numpy as np
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
# 👇 使用 modelscope 库
from transformers import Qwen2_5_VLForConditionalGeneration, AutoTokenizer, AutoProcessor
from qwen_vl_utils import process_vision_info
from transformers import Qwen3VLForConditionalGeneration
import emailgetword
import os
#仅限于使用于2K屏幕远程桌面1080*1920状态
COORDS_FILE = r"F:\AI-OCR\coordinates1.txt"
yanzm = ""
def parse_coordinates(file_path):
    """解析 coordinates.txt，支持无空格格式，如 '名称:LT(x,y),RT(x,y),RB(x,y),LB(x,y)'"""
    coords = {}
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"坐标文件不存在: {file_path}")

    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # 更健壮的正则：允许 : 后无空格，括号不转义错误，逗号前后可有空格
    pattern = re.compile(
        r'^(.*?):\s*'                      # 名称:
        r'LT\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)\s*,\s*'  # LT(x, y),
        r'RT\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)\s*,\s*'  # RT(x, y),
        r'RB\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)\s*,\s*'  # RB(x, y),
        r'LB\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)\s*$'     # LB(x, y)
    )

    for line_num, line in enumerate(lines, 1):
        line = line.strip()
        if not line or line.startswith('#'):
            continue

        match = pattern.match(line)
        if not match:
            print(f"⚠️ 第 {line_num} 行格式错误，跳过: {line}")
            continue

        name = match.group(1).strip()
        # 提取 8 个数字：LT_x, LT_y, RT_x, RT_y, RB_x, RB_y, LB_x, LB_y
        coords_tuple = tuple(int(match.group(i)) for i in range(2, 10))
        lt = (coords_tuple[0], coords_tuple[1])
        rt = (coords_tuple[2], coords_tuple[3])
        rb = (coords_tuple[4], coords_tuple[5])
        lb = (coords_tuple[6], coords_tuple[7])

        # 计算中心点（LT 和 RB 中点）
        center = ((lt[0] + rb[0]) // 2, (lt[1] + rb[1]) // 2)
        rect = (lt[0], lt[1], rb[0], rb[1])  # (x1, y1, x2, y2)

        coords[name] = {
            'LT': lt, 'RT': rt, 'RB': rb, 'LB': lb,
            'center': center,
            'rect': rect
        }
        print(f"✅ 第 {line_num} 行: {name} → center={center}")

    if not coords:
        raise ValueError("未加载任何有效坐标！请检查 coordinates.txt 格式")
    return coords

# 全局加载坐标（启动时一次）
try:
    COORDS = parse_coordinates(COORDS_FILE)
except Exception as e:
    print(f"❌ 坐标加载失败: {e}")
    exit(1)
'''# default: Load the model on the available device(s)
model = Qwen3VLForConditionalGeneration.from_pretrained(
    "Qwen/Qwen3-VL-8B-Instruct", dtype="auto", device_map="auto"
)'''
# =========================
# 配置区域
# =========================
captcha_pil_test = Image.open(r"F:\AI-OCR\captcha_samples\20250819_001558_753202_0014.png")
# =========================
# ✅ 新增配置：表格与搜索相关坐标
# =========================
EXCEL_PATH = r"F:\AI-OCR\改密码清单.xlsx"  # 修改为你的 Excel 路径
# =========================
# =========================
# ✅ 全局短信验证码缓存（批量复用核心）
# =========================
GLOBAL_SMS_CODE_CACHE = "yanzm"  # 初始为空

class SMSError(Exception):
    """短信验证码错误，需刷新缓存"""
    pass
# 初始化 Qwen-VL 模型（从本地加载）
# =========================
def initialize_qwen_vl():
    # 👇 指向你本地的模型文件夹
    model_id = r"F:\AI-OCR\mox\Qwen3-VL-4B-Instruct"

    # 加载模型
    model = Qwen3VLForConditionalGeneration.from_pretrained(
        model_id, torch_dtype="auto", device_map="auto"
    )

    # 加载分词器和处理器
    processor = AutoProcessor.from_pretrained(model_id)
    #processor = AutoProcessor.from_pretrained("Qwen/Qwen2.5-VL-3B-Instruct")

    return model, processor

# 全局初始化
print("🚀 正在加载本地Qwen3-VL-4B-Instruct模型...")
model, processor = initialize_qwen_vl()
print("✅ 模型加载完成！")
# ========================
# OCR 识别主函数（使用 Qwen-VL）
# =========================
def recognize_captcha(x1,y1,x2,y2, text):
    print("📸 正在截取屏幕...")
    screenshot_pil = pyautogui.screenshot()
    # 裁剪出验证码区域
    captcha_pil = screenshot_pil.crop((x1, y1, x2, y2))

    # 保存用于调试
    captcha_pil.save("debug_captcha_qwen.png")

    try:
        # 👇 构建与示例代码完全一致的 messages 结构
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "image", "image": captcha_pil},
                    {"type": "text", "text": text}
                ]
            }
        ]

        # 👇 使用示例代码中的标准流程
        # Preparation for inference
        text = processor.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True
        )
        image_inputs, video_inputs = process_vision_info(messages)
        inputs = processor(
            text=[text],
            images=image_inputs,
            videos=video_inputs,
            padding=True,
            return_tensors="pt",
        )
        # 将输入移动到模型所在的设备
        inputs = inputs.to(model.device)

        # Inference: Generation of the output
        generated_ids = model.generate(**inputs, max_new_tokens=10, do_sample=False)
        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text_list = processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )

        # 通常输出是一个列表，取第一个元素
        response = output_text_list[0] if output_text_list else ""
        # 清理响应
        print(f"🤖 Qwen-VL 识别结果: '{response}'")
        return response

    except Exception as e:
        print(f"❌ Qwen-VL 识别失败: {e}")
        return ""
# =========================
# 主自动化流程
# =========================
# =========================
# 主自动化流程（✅ 按最新21步重写，完全兼容原框架）
# =========================
def main(sms_code_cache="yanzm"):
    global GLOBAL_SMS_CODE_CACHE
    """
    支持复用短信验证码，所有坐标从 COORDS 动态加载
    返回: ("y", code) 成功 / ("x", None) 失败 / raise SMSError 短信错
    """
    print("🎯 全自动密码修改流程（动态坐标 + 短信复用）")
    time.sleep(1)

    # ===== Step 6: 拖选“定期修改密码”文字 =====
    reason_text_lt = COORDS["定期修改密码文字"]['LT']
    reason_text_rb = COORDS["定期修改密码文字"]['RB']
    print(f"🖱️ 拖选‘定期修改密码’区域: LT{reason_text_lt} → RB{reason_text_rb}")
    pyautogui.moveTo(*reason_text_lt)
    time.sleep(0.2)
    pyautogui.dragTo(*reason_text_rb, duration=0.3, button='left')
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    # ===== Step 7: 粘贴到原因输入框 =====
    reason_box_center = COORDS["密码修改原因输入框"]['center']
    print(f"✅ 点击原因输入框 {reason_box_center} → 粘贴")
    pyautogui.click(*reason_box_center)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.5)

    # ===== Step 8–9: 密码输入 =====
    print("✅ 输入新密码与确认密码...")
    pyautogui.click(*COORDS["新密码输入框"]['center']);
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'a');
    time.sleep(0.2);
    pyautogui.typewrite("WangHL@1706");
    time.sleep(0.5)

    pyautogui.click(*COORDS["确认密码输入框"]['center']);
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'a');
    time.sleep(0.2);
    pyautogui.typewrite("WangHL@1706");
    time.sleep(0.5)

    # ===== 预定义坐标（从 COORDS 提取）=====
    SMS_BOX_CENTER = COORDS["短信验证码输入框"]['center']
    GRAPH_INPUT_CENTER = COORDS["图形验证码输入框"]['center']
    GRAPH_IMG_RECT = COORDS["图像验证码图片"]['rect']  # (x1, y1, x2, y2)
    POPUP_RECT = COORDS["弹窗区域"]['rect']  # (x1, y1, x2, y2)
    CLOSE_CENTER = COORDS["弹窗关闭按钮"]['center']
    CHANGE_CAPTCHA_CENTER = COORDS["换一张按钮"]['center']
    CONFIRM_CENTER = COORDS["确认按钮"]['center']
    #code = "yanzm"

    def _fill_sms_code(code):
        if code == "MANUAL":
            return
        pyautogui.click(*SMS_BOX_CENTER);
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'a');
        time.sleep(0.2)
        pyautogui.typewrite(str(code));
        time.sleep(0.5)

    # ===== 首次发送 or 复用缓存 =====
    current_sms_code = sms_code_cache
    _fill_sms_code(current_sms_code)

    # ===== 图形验证码循环 =====
    POPUP_X1, POPUP_Y1, POPUP_X2, POPUP_Y2 = POPUP_RECT
    for graph_retry in range(3):
        print(f"\n🔄 图形码尝试 #{graph_retry + 1}")

        # Step 17: 识别图形码
        captcha_raw = recognize_captcha(
            *GRAPH_IMG_RECT,
            "识别图中验证码：5位大写字母与数字组合。仅返回纯文本。"
        )
        captcha_clean = re.sub(r'[^A-Z0-9]', '', captcha_raw.upper())[:5]
        print(f"🤖 图形码: '{captcha_clean}'")

        if len(captcha_clean) >= 5:
            # Step 18–19: 输入 + 确认
            pyautogui.click(*GRAPH_INPUT_CENTER);
            time.sleep(0.3)
            pyautogui.hotkey('ctrl', 'a');
            time.sleep(0.2)
            pyautogui.typewrite(captcha_clean);
            time.sleep(0.5)
            pyautogui.click(*CONFIRM_CENTER);
            time.sleep(5)

        # === Step 20: 检查弹窗 ===
        if np.array(pyautogui.screenshot().crop(POPUP_RECT)).std() > 5:
            popup_content = recognize_captcha(
                POPUP_X1, POPUP_Y1, POPUP_X2, POPUP_Y2,
                "分析弹窗：\n"
                "- 含‘短信验证码’错误 → sms_error\n"
                "- 含‘验证码’错误但无‘短信’ → captcha_error\n"
                "- 含‘xm3yangchangjing成功’ → success\n"
                "仅回答：sms_error / captcha_error / success"
            ).strip().lower()
            print(f"🔍 弹窗判定: '{popup_content}'")

            if popup_content == "success":
                print("🎉 成功！")
                #这里这个按钮需要识别一下弹窗框的关闭按钮在哪,并返回坐标
                pyautogui.click(*CLOSE_CENTER);
                time.sleep(1)
                return "y", current_sms_code  # ✅ 返回有效码

            elif popup_content == "sms_error":
                print("🔴 短信验证码错误！")
                pyautogui.click(*CLOSE_CENTER);
                time.sleep(0.5)
                raise SMSError("短信验证码失效")

            elif popup_content == "captcha_error":
                print("🟠 图形验证码错误")
                pyautogui.click(*CLOSE_CENTER);
                time.sleep(0.5)
                if graph_retry < 2:
                    pyautogui.click(*CHANGE_CAPTCHA_CENTER);
                    time.sleep(1.5)

        else:
            print("⚠️ 未检测到弹窗，可能超时")
            return "x", None

    return "x", None
def search_and_login(ne_name):
    """
    🔍 全OCR验证版：搜索 → 等5秒 → 用 recognize_captcha 验证按钮状态 → 点击 → 验证跳转
    所有坐标从 COORDS 动态加载
    """

    print(f"🔍 正在搜索网元: {ne_name}")
    my_box_center = COORDS["我的从账号"]['center']
    pyautogui.click(*my_box_center)
    time.sleep(5)
    # === Step 1: 输入设备名 ===
    search_box_center = COORDS["搜索输入框"]['center']
    pyautogui.click(*search_box_center)
    time.sleep(0.5)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.2)
    pyautogui.press('backspace')
    time.sleep(0.2)
    pyautogui.typewrite(ne_name, interval=0.1)
    time.sleep(0.5)

    # === Step 2: 点击搜索 ===
    search_btn_center = COORDS["搜索按钮"]['center']
    pyautogui.click(*search_btn_center)
    print(f"🖱️ 已点击搜索按钮 {search_btn_center}，等待5秒...")
    time.sleep(5)

    # === Step 3: OCR 验证“修改密码”按钮状态 ===
    SCREEN_X1, SCREEN_Y1 = 0, 0
    SCREEN_X2, SCREEN_Y2 = 1920, 642
    screenshot = pyautogui.screenshot()
    result_region = screenshot.crop((SCREEN_X1, SCREEN_Y1, SCREEN_X2, SCREEN_Y2))
    result_region.save(f"debug_search_{ne_name}.png")

    print("🔍 调用 Qwen-VL 验证按钮状态...")
    verdict = recognize_captcha(
        SCREEN_X1, SCREEN_Y1, SCREEN_X2, SCREEN_Y2,
        "请判断：图中是否 **有且仅有一个‘修改密码’按钮**，且无‘暂无数据’‘加载失败’等错误提示？\n"
        "是 → only_one\n否 → no\n仅回答：only_one 或 no"
    ).strip().lower()

    print(f"🤖 按钮状态判定: '{verdict}'")
    if verdict != "only_one":
        print("❌ 按钮验证失败，跳过该网元")
        return False

    # === Step 4: 点击“修改密码”按钮 ===
    modify_btn_center = COORDS["修改密码按钮"]['center']
    print(f"⚙️ 点击‘修改密码’按钮 {modify_btn_center}...")
    pyautogui.click(*modify_btn_center)
    time.sleep(1)

    # === Step 5: 等待跳转 + OCR 验证页面 ===
    print("⏳ 等待页面跳转（8秒）...")
    time.sleep(8)

    # 验证区域：覆盖“定期修改密码”提示 + 原因输入框
    VERIFY_RECT = (350, 400, 800, 500)
    print("🔍 OCR验证：是否已进入密码修改页面？")
    page_check = recognize_captcha(
        *VERIFY_RECT,
        "图中是否为密码修改页面？检查：\n"
        "- 有‘密码修改原因’输入框？\n"
        "- 有‘定期修改密码’提示文字？\n"
        "- 无‘搜索’‘登录’等首页元素？\n"
        "是 → yes\n否 → no\n仅回答：yes 或 no"
    ).strip().lower()

    if page_check == "yes":
        print("✅ 跳转验证通过：已进入密码修改页面")
        return True
    else:
        print("❌ 跳转失败：未检测到密码修改页面")
        return False
# =========================
# ✅ 新增主流程：处理 Excel 并逐行执行
def _update_excel_status(excel_path, row, status):
    """安全更新 Excel 某行第2列状态"""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=row, column=2, value=status)
        wb.save(excel_path)
        wb.close()
        print(f"💾 已更新第 {row} 行状态为: '{status}'")
    except Exception as e:
        print(f"⚠️ 更新状态失败（第 {row} 行）: {e}")
def main_with_excel():
    global GLOBAL_SMS_CODE_CACHE
    time.sleep(1)
    print("📋 正在读取 Excel 表格...")
    try:
        df = pd.read_excel(EXCEL_PATH)
    except Exception as e:
        print(f"❌ 无法读取 Excel 文件: {e}")
        return

    # 处理列
    ne_column = df.columns[0]
    status_column = df.columns[1] if len(df.columns) > 1 else "状态"
    if len(df.columns) < 2:
        df[status_column] = ""

    records = []
    for idx, row in df.iterrows():
        ne_name = str(row[ne_column]).strip()
        status = str(row[status_column]).strip() if not pd.isna(row[status_column]) else ""
        if ne_name and ne_name.lower() != 'nan':
            records.append((idx, ne_name, status))

    print(f"✅ 共读取到 {len(records)} 个网元")

    for idx_in_df, ne_name, current_status in records:
        excel_row = idx_in_df + 2
        print(f"\n📌 处理第 {excel_row-1}/{len(records)} 个网元: {ne_name} | 状态: '{current_status}'")

        if current_status == "是":
            print("⏭️ 已完成，跳过")
            continue

        try:
            if not search_and_login(ne_name):
                print("⏭️ 搜索登录失败")
                _update_excel_status(EXCEL_PATH, excel_row, "否")
                continue
            time.sleep(2)

            # ===== 核心：复用短信码 + 智能刷新 =====
            success = False
            device_retry = 0
            while device_retry < 2 and not success:  # 每设备最多2次尝试（含刷新）
                try:
                    result, used_code = main(sms_code_cache=GLOBAL_SMS_CODE_CACHE)  # ✅ 接收返回码
                    if result == "y":
                        success = True
                        # ✅ 首次成功 or 码更新时，缓存最新有效码
                        if used_code and used_code != "MANUAL":
                            GLOBAL_SMS_CODE_CACHE = used_code
                            print(f"🆕 全局短信码已更新: {used_code}")
                except SMSError:
                    # 刷新缓存
                    print("🔧 短信码失效，正在刷新最新验证码...")
                    GLOBAL_SMS_CODE_CACHE = yanzm
                    device_retry += 1
                    continue
                except Exception as e:
                    print(f"❌ main() 异常: {e}")
                    break

                if not success:
                    device_retry += 1
                    if device_retry < 2:
                        print(f"🔁 设备重试第 {device_retry} 次...")
                        time.sleep(3)

            if success:
                print(f"✅ {ne_name} 成功")
                _update_excel_status(EXCEL_PATH, excel_row, "是")
                # 首次成功时缓存码（后续设备复用）
                if GLOBAL_SMS_CODE_CACHE is None:
                    # 实际码已在 main() 中填入，但未返回 → 可通过邮箱再取一次，或留空（后续出错再刷新）
                    # 此处暂不主动缓存，依赖首次 main() 内部获取
                    pass
            else:
                print(f"❌ {ne_name} 失败")
                _update_excel_status(EXCEL_PATH, excel_row, "否")

            print("🔙 返回上一页...")
            pyautogui.press('esc')
            time.sleep(2)

        except Exception as e:
            print(f"❌ 处理 {ne_name} 出错: {e}")
            _update_excel_status(EXCEL_PATH, excel_row, f"异常")
            continue

    print("🎉 批量处理完成！")

# =========================
# 启动脚本
# =========================
if __name__ == "__main__":
    print("🚀 全自动密码修改脚本已启动")
    print("📦 正在加载模型（仅一次）...")

    # ===== 1. 全局加载模型 =====
    #model, processor = initialize_qwen_vl()
    #print("✅ 模型加载完成！")

    # ===== 2. 循环主菜单 =====
    while True:
        print("\n" + "=" * 50)
        print("🔄 当前模式：支持多次连续执行")
        choice = input(
            "请选择操作：\n"
            "  🟢 输入 y → 执行一次密码修改（请确保已切换到目标页面）\n"
            "  🔵 输入 e → 启动 Excel 批量处理（处理完继续循环）\n"
            "  🔴 输入 q → 退出程序\n"
            "请输入: "
        ).strip().lower()

        if choice == 'y':
            print("🎯 开始执行单次密码修改...")
            print("👉 系统将在 3 秒后开始操作，请确保窗口已就绪！")
            time.sleep(3)  # 给你时间切换或准备

            try:
                result = main()  # 调用你的 main 函数
                if result == "y":
                    print("🎉 本次密码修改成功！")
                elif result == "x":
                    print("⚠️ 本次修改失败（验证码错误等），可重试。")
                else:
                    print("❌ 本次执行异常或超时。")
            except Exception as e:
                print(f"❌ 执行 main() 时发生异常: {e}")

            # ✅ 不退出，继续循环
            input("↩️ 单次执行结束，按 Enter 键返回主菜单...")

        elif choice == 'e':
            print("📊 即将启动 Excel 批量处理...")
            print("👉 系统将在 3 秒后开始操作，请确保窗口已就绪！")

            confirm = input("👉 请确保 Excel 文件已关闭！确认开始？(y/n): ").strip().lower()
            if confirm == 'y':
                try:
                    time.sleep(5)  # 给你时间切换或准备
                    main_with_excel()
                except Exception as e:
                    print(f"❌ 批量处理出错: {e}")
            else:
                print("⏭️ 已取消批量处理。")
            input("↩️ 批量处理结束，按 Enter 键返回主菜单...")

        elif choice == 'q':
            print("👋 感谢使用，再见！")
            break  # 退出循环

        else:
            print("❌ 输入无效，请输入 y、e 或 q。")
            time.sleep(1)
