import pyautogui
import time
import re
import numpy as np
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from transformers import Qwen3VLForConditionalGeneration, AutoProcessor
from qwen_vl_utils import process_vision_info
import emailgetword
import random
import os

# ========================
# 全局变量 & 初始化
# ========================
model = None
processor = None
#GLOBAL_SMS_CODE_CACHE = None
GLOBAL_SMS_CODE_CACHE = None
GLOBAL_SEARCH_PAGE_ELEMENTS = None   # 👈 新增
GLOBAL_PASSWORD_PAGE_ELEMENTS = None # 👈 新增
import pygetwindow as gw

def get_remote_desktop_window_rect():
    """
    通过系统 API 获取“远程桌面连接”窗口的位置
    返回: (x, y, width, height) 或 None
    """
    try:
        # 查找标题包含“远程桌面连接”的窗口
        windows = gw.getWindowsWithTitle('远程桌面连接')
        if not windows:
            # 尝试英文标题
            windows = gw.getWindowsWithTitle('Remote Desktop Connection')
        if not windows:
            print("❌ 未找到远程桌面窗口")
            return None

        # 取第一个匹配的窗口
        win = windows[0]
        if win.isMinimized:
            win.restore()  # 恢复最小化窗口
        x, y = win.left, win.top
        width, height = win.width, win.height
        print(f"✅ 远程桌面窗口: x={x}, y={y}, w={width}, h={height}")
        return (x, y, width, height)
    except Exception as e:
        print(f"⚠️ 获取窗口失败: {e}")
        return None

class SMSError(Exception):
    pass

def initialize_qwen_vl():
    model_id = r"F:\AI-OCR\mox\Qwen3-VL-4B-Instruct"
    model = Qwen3VLForConditionalGeneration.from_pretrained(
        model_id, torch_dtype="auto", device_map="auto"
    )
    processor = AutoProcessor.from_pretrained(model_id)
    return model, processor

print("🚀 正在加载本地Qwen3-VL-4B-Instruct模型...")
model, processor = initialize_qwen_vl()
print("✅ 模型加载完成！")

# ========================
# OCR 识别主函数（使用 Qwen-VL）
# =========================
def recognize_captcha(x1, y1, x2, y2, text):
    print("📸 正在截取屏幕...")
    screenshot_pil = pyautogui.screenshot()
    captcha_pil = screenshot_pil.crop((x1, y1, x2, y2))
    captcha_pil.save("debug_captcha_qwen.png")

    try:
        messages = [{
            "role": "user",
            "content": [
                {"type": "image", "image": captcha_pil},
                {"type": "text", "text": text}
            ]
        }]

        text_prompt = processor.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
        image_inputs, _ = process_vision_info(messages)
        inputs = processor(
            text=[text_prompt],
            images=image_inputs,
            padding=True,
            return_tensors="pt",
        ).to(model.device)

        generated_ids = model.generate(**inputs, max_new_tokens=10, do_sample=False)
        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text_list = processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        response = output_text_list[0] if output_text_list else ""
        print(f"🤖 Qwen-VL 识别结果: '{response}'")
        return response

    except Exception as e:
        print(f"❌ Qwen-VL 识别失败: {e}")
        return ""

# ========================
# 动态定位工具函数
# ========================
def unprocess_bbox_by_ratio(bbox, screen_size):
    x1, y1, x2, y2 = bbox
    W_screen, H_screen = screen_size
    W_model, H_model = 1000, 1000
    X1 = int(round(x1 * W_screen / W_model))
    Y1 = int(round(y1 * H_screen / H_model))
    X2 = int(round(x2 * W_screen / W_model))
    Y2 = int(round(y2 * H_screen / H_model))
    X1 = max(0, min(W_screen, X1))
    Y1 = max(0, min(H_screen, Y1))
    X2 = max(0, min(W_screen, X2))
    Y2 = max(0, min(H_screen, Y2))
    return X1, Y1, X2, Y2

def detect_element(image_pil, description):
    prompt = f"""# 视觉定位
    请检测图中 {description} 的边界框，并严格输出：
    <ref>{description}</ref><box>(x1,y1),(x2,y2)</box>"""

    try:
        messages = [{
            "role": "user",
            "content": [
                {"type": "image", "image": image_pil},
                {"type": "text", "text": prompt}
            ]
        }]

        text = processor.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
        image_inputs, _ = process_vision_info(messages)
        inputs = processor(text=[text], images=image_inputs, return_tensors="pt").to(model.device)

        generated_ids = model.generate(**inputs, max_new_tokens=150, do_sample=False)
        output = processor.decode(generated_ids[0][len(inputs.input_ids[0]):], skip_special_tokens=True).strip()
        print(f"🔍 原始输出: '{output}'")

        match_bbox = re.search(r"<ref>.*?</ref><box>\s*\((\d+),(\d+)\)\s*,\s*\((\d+),(\d+)\)\s*</box>", output)
        if not match_bbox:
            print("⚠️ 未找到目标 bbox")
            return None

        bbox = tuple(map(int, match_bbox.groups()))
        W_screen, H_screen = image_pil.size
        X1, Y1, X2, Y2 = unprocess_bbox_by_ratio(bbox, (W_screen, H_screen))
        return X1, Y1, X2, Y2

    except Exception as e:
        print(f"❌ detect_element 错误: {e}")
        import traceback
        traceback.print_exc()
        return None

def split_screen_into_quadrants(screen_width=1920, screen_height=1080):
    x,y,w, h = 0,0,screen_width, screen_height
    #x,y,w,h = get_remote_desktop_window_rect()
    half_w, half_h = int(w // 2), int(h // 2)
    return [
        (x, y, half_w, half_h),           # 左上
        (half_w, y, w, half_h),            # 右上
        (x, half_h, half_w, h),             # 左下
        (half_w, half_h, w, h)              # 右下
    ]

'''def detect_element_global(description):
    """常规元素使用四等分策略"""
    screen = pyautogui.screenshot()
    quadrants = split_screen_into_quadrants()

    for i, (qx1, qy1, qx2, qy2) in enumerate(quadrants):
        quadrant_img = screen.crop((qx1, qy1, qx2, qy2))
        bbox_rel = detect_element(quadrant_img, description)
        if bbox_rel:
            rel_x1, rel_y1, rel_x2, rel_y2 = bbox_rel
            abs_x1 = qx1 + rel_x1
            abs_y1 = qy1 + rel_y1
            abs_x2 = qx1 + rel_x2
            abs_y2 = qy1 + rel_y2
            print(f"✅ 在第 {i+1} 象限找到 [{description}] @ ({abs_x1}, {abs_y1}) - ({abs_x2}, {abs_y2})")
            return (abs_x1, abs_y1, abs_x2, abs_y2)

    print(f"❌ 在所有象限中未找到 [{description}]")
    return None'''


def detect_element_global(description):
    """常规元素使用四等分策略，并保存带标注的调试图"""
    screen = pyautogui.screenshot()
    quadrants = split_screen_into_quadrants()

    for i, (qx1, qy1, qx2, qy2) in enumerate(quadrants):
        quadrant_img = screen.crop((qx1, qy1, qx2, qy2))

        # 👇 新增：在象限图上叠加红色边框和编号
        from PIL import ImageDraw
        draw = ImageDraw.Draw(quadrant_img)
        # 画红色边框
        draw.rectangle([0, 0, quadrant_img.width - 1, quadrant_img.height - 1], outline="red", width=3)
        # 写象限编号（左上角）
        draw.text((5, 5), f"Quadrant {i}", fill="red")
        # 保存带标注的图
        quadrant_img.save(f"debug_quadrant_{i}.png")

        print(f"🔍 正在检测第 {i + 1} 象限...")
        bbox_rel = detect_element(quadrant_img, description)
        if bbox_rel:
            rel_x1, rel_y1, rel_x2, rel_y2 = bbox_rel
            abs_x1 = qx1 + rel_x1
            abs_y1 = qy1 + rel_y1
            abs_x2 = qx1 + rel_x2
            abs_y2 = qy1 + rel_y2
            print(f"✅ 在第 {i + 1} 象限找到 [{description}] @ ({abs_x1}, {abs_y1}) - ({abs_x2}, {abs_y2})")
            return (abs_x1, abs_y1, abs_x2, abs_y2)

    print(f"❌ 在所有象限中未找到 [{description}]")
    return None
def get_center(bbox):
    x1, y1, x2, y2 = bbox
    cx = (x1 + x2) // 2
    cy = (y1 + y2) // 2
    cx += random.randint(-2, 2)
    cy += random.randint(-2, 2)
    return cx, cy

# ========================
# 弹窗专用 ROI 检测（避免四等分切割）
# ========================


# 删除原 def detect_popup_region(): 和 def detect_popup_close_button():

# 新增：全局缓存远程桌面窗口
GLOBAL_REMOTE_DESKTOP_WINDOW = None

def detect_popup_region():
    """
    使用系统 API 获取远程桌面窗口，以窗口中心点为基准，±250px 计算弹窗 ROI
    返回: (x1, y1, x2, y2) 全局坐标
    """
    rect = get_remote_desktop_window_rect()
    if not rect:
        print("⚠️ 无法获取远程桌面窗口，使用默认 ROI (730,450,1200,750)")
        return (730, 450, 1200, 750)

    win_x, win_y, win_w, win_h = rect
    # 计算窗口中心点
    center_x = win_x + win_w // 2
    center_y = win_y + win_h // 2

    # 弹窗 ROI：中心点 ±250 像素
    popup_x1 = max(win_x, center_x - 250)
    popup_y1 = max(win_y, center_y - 250)
    popup_x2 = min(win_x + win_w, center_x + 250)
    popup_y2 = min(win_y + win_h, center_y + 250)

    print(f"🖼️ 动态弹窗 ROI (基于窗口中心): ({popup_x1}, {popup_y1}) - ({popup_x2}, {popup_y2})")
    return (popup_x1, popup_y1, popup_x2, popup_y2)


def detect_popup_close_button():
    """在动态 ROI 内识别关闭按钮"""
    popup_roi = detect_popup_region()
    screen = pyautogui.screenshot()
    popup_img = screen.crop(popup_roi)
    popup_img.save("debug_popup_roi.png")

    bbox_rel = detect_element(popup_img, "弹窗右上角或右下角的‘×’关闭图标")
    if bbox_rel:
        rel_x1, rel_y1, rel_x2, rel_y2 = bbox_rel
        abs_x1 = popup_roi[0] + rel_x1
        abs_y1 = popup_roi[1] + rel_y1
        abs_x2 = popup_roi[0] + rel_x2
        abs_y2 = popup_roi[1] + rel_y2
        return (abs_x1, abs_y1, abs_x2, abs_y2)
    else:
        return None


def close_popup(popup_bbox):
    close_bbox = detect_popup_close_button()
    if close_bbox:
        cx, cy = get_center(close_bbox)
        print(f"✅ 动态识别到关闭按钮，中心点: ({cx}, {cy})")
        pyautogui.click(cx, cy)
    else:
        print("⚠️ 未识别到关闭按钮，点击弹窗中心关闭")
        pyautogui.click(*get_center(popup_bbox))
    time.sleep(0.5)

# ========================
# 主自动化流程（预存坐标模式）
# ========================
ELEMENTS_TO_DETECT = [
    ("periodic_text", "‘定期修改密码’文字（位于密码修改原因输入框旁边）"),
    ("reason_input", "密码修改原因输入框（位于‘密码修改原因’标签右侧）"),
    ("new_pwd_input", "新密码输入框（位于‘新密码’标签右侧）"),
    ("confirm_pwd_input", "确认密码输入框（位于‘确认密码’标签右侧）"),
    ("sms_input", "短信验证码输入框（位于‘短信验证码’标签右侧）"),
    ("captcha_img", "位于‘验证码’输入框右侧的5位字母数字验证码图片"),
    ("captcha_input", "验证码输入框（位于验证码图片左侧）"),
    ("change_captcha", "‘换一张’链接（位于验证码图片右侧）"),
    ("confirm_btn", "蓝色的‘确认’按钮"),
]

def detect_all_elements_once(element_descriptions):
    elements = {}
    for name, desc in element_descriptions:
        print(f"🔍 正在识别 [{name}] ...")
        bbox = detect_element_global(desc)
        if bbox:
            elements[name] = bbox
            print(f"✅ [{name}] → bbox={bbox}")
        else:
            print(f"⚠️ [{name}] 未找到")
    return elements
def debug_captcha_region(x1, y1, x2, y2):
    """保存验证码区域截图，用于人工检查"""
    screen = pyautogui.screenshot()
    captcha_img = screen.crop((x1, y1, x2, y2))
    captcha_img.save("debug_captcha_raw.png")
    # 转灰度 + 二值化
    captcha_img = captcha_img.convert('L')
    captcha_img = captcha_img.point(lambda x: 0 if x < 128 else 255, '1')
    captcha_img.save("debug_captcha_processed.png")
    print(f"✅ 已保存验证码截图，请检查 debug_captcha_*.png")

def main(sms_code_cache=None):
    global GLOBAL_SMS_CODE_CACHE, GLOBAL_PASSWORD_PAGE_ELEMENTS
    print("🎯 全自动密码修改流程（全局坐标缓存模式）")
    time.sleep(1)

    # === 首次进入密码页：识别并缓存 ===
    if GLOBAL_PASSWORD_PAGE_ELEMENTS is None:
        print("🖼️ 首次进入密码页，正在预存所有元素坐标...")

        # Step 1: 先点短信框触发验证码
        sms_input_bbox = detect_element_global("短信验证码输入框（位于‘短信验证码’标签右侧）")
        if not sms_input_bbox:
            print("❌ 未找到短信验证码输入框")
            return "x", None
        SMS_BOX_CENTER = get_center(sms_input_bbox)
        pyautogui.click(*SMS_BOX_CENTER)
        time.sleep(0.5)
        # 关闭浏览器缓存提示
        label_x = sms_input_bbox[0] - 80
        label_y = (sms_input_bbox[1] + sms_input_bbox[3]) // 2
        pyautogui.click(label_x, label_y)
        time.sleep(1.5)  # 等待验证码加载

        # Step 2: 识别所有元素
        STATIC_ELEMENTS = [
            ("periodic_text", "‘ 定期修改密码 ’文字（位于密码修改原因输入框旁边）"),
            ("reason_input", "密码修改原因输入框（位于‘密码修改原因’标签右侧）"),
            ("new_pwd_input", "新密码输入框（位于‘新密码’标签右侧）"),
            ("confirm_pwd_input", "确认密码输入框（位于‘确认密码’标签右侧）"),
            ("sms_input", "短信验证码输入框（位于‘短信验证码’标签右侧）"),
            ("confirm_btn", "蓝色的‘确认’按钮"),
            ("captcha_input", "验证码输入框（位于验证码图片左侧）"),
            ("change_captcha", "‘换一张’链接（位于验证码图片右侧）"),
            ("captcha_img", "位于‘验证码’输入框右侧的5位字母数字验证码图片"),
        ]
        DYNAMIC_ELEMENTS = [

        ]

        GLOBAL_PASSWORD_PAGE_ELEMENTS = {}
        # 静态元素
        static_elems = detect_all_elements_once(STATIC_ELEMENTS)
        GLOBAL_PASSWORD_PAGE_ELEMENTS.update(static_elems)
        # 动态元素
        dynamic_elems = detect_all_elements_once(DYNAMIC_ELEMENTS)
        GLOBAL_PASSWORD_PAGE_ELEMENTS.update(dynamic_elems)

        # 加入短信框
        GLOBAL_PASSWORD_PAGE_ELEMENTS["sms_input"] = sms_input_bbox

    # 后续所有操作直接使用 GLOBAL_PASSWORD_PAGE_ELEMENTS
    page_elements = GLOBAL_PASSWORD_PAGE_ELEMENTS

    required_static = ["periodic_text", "reason_input", "new_pwd_input", "confirm_pwd_input", "sms_input",
                       "confirm_btn","captcha_input", "change_captcha"]
    for r in required_static:
        if r not in page_elements:
            print(f"❌ 静态元素 [{r}] 缺失，无法继续")
            return "x", None

    # === Step 2: 执行密码相关操作（按原有顺序）===
    # 拖选“定期修改密码”文字
    lt_x, lt_y, rb_x, rb_y = page_elements["periodic_text"]
    pyautogui.moveTo(rb_x, rb_y)
    time.sleep(0.2)
    pyautogui.dragTo(lt_x, lt_y, duration=0.3, button='left')
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.5)

    # 粘贴到原因输入框
    cx, cy = get_center(page_elements["reason_input"])
    pyautogui.click(cx, cy)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.2)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.5)

    # 输入密码
    def _type_at_bbox(bbox, text):
        cx, cy = get_center(bbox)
        pyautogui.click(cx, cy)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.2)
        pyautogui.typewrite(text, interval=0.05)
        time.sleep(0.5)

    _type_at_bbox(page_elements["new_pwd_input"], "WangHL@1706")
    _type_at_bbox(page_elements["confirm_pwd_input"], "WangHL@1706")

    # === Step 3: 点击短信验证码输入框，触发图形验证码加载 ===
    sms_bbox = page_elements["sms_input"]
    SMS_BOX_CENTER = get_center(sms_bbox)
    pyautogui.click(*SMS_BOX_CENTER)
    print(f"✅ 已点击短信验证码输入框 @ {SMS_BOX_CENTER}")
    def _ensure_no_suggestions():
        """点击屏幕右下角空白区域，确保无浮动框"""
        screen_width, screen_height = pyautogui.size()
        pyautogui.click(100, 700)
        time.sleep(0.3)

    _ensure_no_suggestions()
    time.sleep(2)  # ⚠️ 等待图形验证码加载

    # === Step 4: 识别动态元素（图形验证码相关）===
    DYNAMIC_ELEMENTS = [
        ("captcha_img", "位于‘验证码’输入框右侧的5位字母数字验证码图片"),
    ]
    print("🖼️ 正在识别动态元素（验证码相关）...")
    dynamic_elements = detect_all_elements_once(DYNAMIC_ELEMENTS)

    # 合并到 page_elements
    page_elements.update(dynamic_elements)

    required_dynamic = ["captcha_img"]
    for r in required_dynamic:
        if r not in page_elements:
            print(f"❌ 动态元素 [{r}] 缺失，可能验证码未加载")
            return "x", None
    # === Step 5: 短信验证码处理（复用你的逻辑）===
    def _trigger_sms_send():
        pyautogui.click(*SMS_BOX_CENTER)
        time.sleep(0.3)
        sms_x1, _, sms_x2, _ = page_elements["sms_input"]
        GET_SMS_BTN = (sms_x2 + 20, SMS_BOX_CENTER[1])
        pyautogui.click(*GET_SMS_BTN)
        time.sleep(0.5)

        for _ in range(8):
            time.sleep(1)
            popup_bbox = detect_popup_region()
            if popup_bbox:
                x1, y1, x2, y2 = popup_bbox
                txt = recognize_captcha(x1, y1, x2, y2, "是否含‘发送成功’？yes/no")
                if "yes" in txt.lower():
                    close_popup(popup_bbox)
                    return True
        return False

    def _get_latest_sms_code():
        try:
            for attempt in range(12):
                code = emailgetword.get_verification_code_once(check_limit=10)
                if code and len(code) >= 4:
                    print(f"📬 获取最新验证码: {code}")
                    return code
                time.sleep(5)
        except Exception as e:
            print(f"⚠️ 邮箱模块异常: {e}")
        input("✅ 请手动输入最新短信码后按 Enter...")
        return "MANUAL"

    def _fill_sms_code(code):
        if code == "MANUAL":
            return
        pyautogui.click(*SMS_BOX_CENTER)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.2)
        pyautogui.typewrite(str(code))
        time.sleep(0.5)

    current_sms_code = sms_code_cache
    if sms_code_cache is None:
        print("📱 首次发送短信验证码...")
        if not _trigger_sms_send():
            return "x", None
        time.sleep(60)
        current_sms_code = _get_latest_sms_code()
    else:
        print(f"🔁 复用缓存短信验证码: {sms_code_cache}")

    _fill_sms_code(current_sms_code)

    # === Step 6: 图形验证码循环 ===
    for graph_retry in range(3):
        print(f"\n🔄 图形码尝试 #{graph_retry + 1}")

        x1, y1, x2, y2 = page_elements["captcha_img"]
        debug_captcha_region(x1, y1, x2, y2)
        captcha_raw = recognize_captcha(x1, y1, x2, y2, "识别图中验证码：5位大写字母与数字组合。仅返回纯文本。")
        captcha_clean = re.sub(r'[^A-Z0-9]', '', captcha_raw.upper())[:5]
        print(f"🤖 图形码: '{captcha_clean}'")

        if len(captcha_clean) >= 5:
            _type_at_bbox(page_elements["captcha_input"], captcha_clean)
            cx, cy = get_center(page_elements["confirm_btn"])
            pyautogui.click(cx, cy)
            time.sleep(5)

        # 检查弹窗
        popup_bbox = detect_popup_region()
        if popup_bbox:
            x1, y1, x2, y2 = popup_bbox
            popup_content = recognize_captcha(
                x1, y1, x2, y2,
                "分析弹窗：\n"
                "- 先判断是否包含连续的‘修改密码成功’ → success\n"
                "- 包含连续的‘短信验证码错误’错误或连续的'短信校验码未填写'错误 → sms_error\n"
                "- 包含连续的‘验证码错误’但无‘短信’ → captcha_error\n"
                "仅回答：sms_error / captcha_error / success"
            ).strip().lower()
            print(f"🔍 弹窗判定: '{popup_content}'")

            if popup_content == "success":
                print("🎉 成功！正在关闭弹窗...")
                close_popup(popup_bbox)
                return "y", current_sms_code

            elif popup_content == "sms_error":
                print("🔴 短信验证码错误！")
                close_popup(popup_bbox)
                raise SMSError("短信验证码失效")

            elif popup_content == "captcha_error":
                print("🟠 图形验证码错误")
                close_popup(popup_bbox)
                if graph_retry < 2:
                    cx, cy = get_center(page_elements["change_captcha"])
                    pyautogui.click(cx, cy)
                    time.sleep(1.5)
        else:
            print("⚠️ 未检测到弹窗，可能超时")
            return "x", None

    return "x", None
# ========================
# 搜索与登录（保持原逻辑）
# ========================
def search_and_login(ne_name):
    global GLOBAL_SEARCH_PAGE_ELEMENTS
    print(f"🔍 正在搜索网元: {ne_name}")

    # === 首次进入搜索页：识别并缓存 ===
    if GLOBAL_SEARCH_PAGE_ELEMENTS is None:
        print("🖼️ 首次进入搜索页，正在预存所有元素坐标...")
        SEARCH_ELEMENTS = [
            ("my_account", "顶部导航栏中的‘我的从账号’标签"),
            ("search_input", "搜索输入框（位于页面顶部，上面有'从帐号密码批量修改'字符按钮,右边有‘搜索’按钮）"),
            ("search_btn", "蓝色的‘搜索’按钮（位于搜索输入框右侧）"),
            ("modify_pwd_btn", "蓝色的‘修改密码’按钮（位于搜索框下方的第一个）"),
        ]
        GLOBAL_SEARCH_PAGE_ELEMENTS = {}
        for name, desc in SEARCH_ELEMENTS:
            bbox = detect_element_global(desc)
            if bbox:
                GLOBAL_SEARCH_PAGE_ELEMENTS[name] = bbox
                print(f"✅ [{name}] → bbox={bbox}")
            else:
                print(f"⚠️ [{name}] 未找到")

    # 使用缓存坐标操作
    def _click_from_cache(name, fallback_desc=None):
        if name in GLOBAL_SEARCH_PAGE_ELEMENTS:
            cx, cy = get_center(GLOBAL_SEARCH_PAGE_ELEMENTS[name])
            pyautogui.click(cx, cy)
            return True
        elif fallback_desc:
            return locate_and_click(fallback_desc)
        return False

    # 点击“我的从账号”
    _click_from_cache("my_account", "顶部导航栏中的‘我的从账号’标签")
    time.sleep(5)

    # 输入设备名
    if "search_input" in GLOBAL_SEARCH_PAGE_ELEMENTS:
        cx, cy = get_center(GLOBAL_SEARCH_PAGE_ELEMENTS["search_input"])
        pyautogui.click(cx, cy)
        time.sleep(0.3)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.2)
        pyautogui.typewrite(ne_name, interval=0.1)
    else:
        type_text("搜索输入框（位于页面顶部，旁边有‘搜索’按钮）", ne_name)
    time.sleep(0.5)

    # 点击搜索按钮
    _click_from_cache("search_btn", "蓝色的‘搜索’按钮（位于搜索输入框右侧）")
    print("🖱️ 已点击搜索按钮，等待5秒...")
    time.sleep(5)

    # ========================
    # ✅ 关键修复：用 modify_pwd_btn 的四角作为 OCR 区域
    # ========================
    if "modify_pwd_btn" not in GLOBAL_SEARCH_PAGE_ELEMENTS:
        print("❌ 未找到‘修改密码’按钮，跳过该网元")
        return False

    x1, y1, x2, y2 = GLOBAL_SEARCH_PAGE_ELEMENTS["modify_pwd_btn"]

    # 扩展区域以包含上下文（如错误提示）
    def expand_bbox(bbox, margin=80, screen_size=(1920, 1080)):
        x1, y1, x2, y2 = bbox
        w, h = screen_size
        return (
            max(0, x1 - margin),
            max(0, y1 - margin),
            min(w, x2 + margin),
            min(h, y2 + margin)
        )

    ocr_x1, ocr_y1, ocr_x2, ocr_y2 = expand_bbox((x1, y1, x2, y2))

    # 保存 OCR 区域截图（调试用）
    screen = pyautogui.screenshot()
    ocr_region = screen.crop((ocr_x1, ocr_y1, ocr_x2, ocr_y2))
    ocr_region.save("debug_modify_pwd_ocr.png")

    verdict = recognize_captcha(
        ocr_x1, ocr_y1, ocr_x2, ocr_y2,
        "请判断：图中是否 **有且仅有一个‘修改密码’按钮**，且无‘暂无数据’‘加载失败’等错误提示？\n"
        "是 → only_one\n否 → no\n仅回答：only_one 或 no"
    ).strip().lower()

    if verdict != "only_one":
        print("❌ 按钮验证失败，跳过该网元")
        return False

    # 点击“修改密码”按钮
    _click_from_cache("modify_pwd_btn", "‘修改密码’按钮（位于搜索框下方的第一个）")
    time.sleep(1)

    print("⏳ 等待页面跳转（8秒）...")
    time.sleep(8)

    # 验证是否进入密码修改页面
    # ========================
    # ✅ 改进：用四等分全屏识别来判断是否为密码修改页面
    # ========================
    print("🔍 正在验证是否进入密码修改页面...")

    # 用 detect_element_global 检测关键元素
    reason_input_bbox = detect_element_global("密码修改原因输入框（位于‘密码修改原因’标签右侧）")
    # 判断逻辑
    has_reason_input = reason_input_bbox is not None


    if has_reason_input :
        print("✅ 确认为密码修改页面")
        page_check = "yes"
    else:
        print(f"❌ 页面验证失败: 原因输入框={has_reason_input}")
        page_check = "no"

    return page_check == "yes"
# ========================
# 辅助函数（用于 search_and_login）

# ========================
def locate_and_click(description, offset=(0, 0), delay=0.5, jitter=2):
    bbox = detect_element_global(description)
    if bbox:
        x1, y1, x2, y2 = bbox
        cx = (x1 + x2) // 2 + offset[0]
        cy = (y1 + y2) // 2 + offset[1]
        cx += random.randint(-jitter, jitter)
        cy += random.randint(-jitter, jitter)
        print(f"🖱️ 动态点击 [{description}] @ ({cx}, {cy})")
        pyautogui.click(cx, cy)
        time.sleep(delay)
        return True
    else:
        print(f"❌ 无法定位 [{description}]，跳过点击")
        return False

def type_text(description, text, clear_first=True, enter_after=False, delay_char=0.1):
    if not locate_and_click(description):
        return False
    if clear_first:
        pyautogui.hotkey('ctrl', 'a')
        pyautogui.press('delete')
    pyautogui.typewrite(text, interval=delay_char)
    if enter_after:
        pyautogui.press('enter')
    print(f"⌨️ 输入至 [{description}]: '{text}'")
    return True

# ========================
# Excel 批量处理（保持原逻辑）
# ========================
def _update_excel_status(excel_path, row, status):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=row, column=2, value=status)
        wb.save(excel_path)
        wb.close()
        print(f"💾 已更新第 {row} 行状态为: '{status}'")
    except Exception as e:
        print(f"⚠️ 更新状态失败（第 {row} 行）: {e}")

def main_with_excel():
    global GLOBAL_SMS_CODE_CACHE
    EXCEL_PATH = r"F:\AI-OCR\改密码清单.xlsx"
    time.sleep(1)
    print("📋 正在读取 Excel 表格...")
    try:
        df = pd.read_excel(EXCEL_PATH)
    except Exception as e:
        print(f"❌ 无法读取 Excel 文件: {e}")
        return

    ne_column = df.columns[0]
    status_column = df.columns[1] if len(df.columns) > 1 else "状态"
    if len(df.columns) < 2:
        df[status_column] = ""

    records = []
    for idx, row in df.iterrows():
        ne_name = str(row[ne_column]).strip()
        status = str(row[status_column]).strip() if not pd.isna(row[status_column]) else ""
        if ne_name and ne_name.lower() != 'nan':
            records.append((idx, ne_name, status))

    print(f"✅ 共读取到 {len(records)} 个网元")

    for idx_in_df, ne_name, current_status in records:
        excel_row = idx_in_df + 2
        print(f"\n📌 处理第 {excel_row-1}/{len(records)} 个网元: {ne_name} | 状态: '{current_status}'")

        if current_status == "是":
            print("⏭️ 已完成，跳过")
            continue

        try:
            if not search_and_login(ne_name):
                print("⏭️ 搜索登录失败")
                _update_excel_status(EXCEL_PATH, excel_row, "否")
                continue
            time.sleep(2)

            success = False
            device_retry = 0
            while device_retry < 2 and not success:
                try:
                    result, used_code = main(sms_code_cache=GLOBAL_SMS_CODE_CACHE)
                    if result == "y":
                        success = True
                        if used_code and used_code != "MANUAL":
                            GLOBAL_SMS_CODE_CACHE = used_code
                            print(f"🆕 全局短信码已更新: {used_code}")
                except SMSError:
                    print("🔧 短信码失效，正在刷新最新验证码...")
                    locate_and_click("短信验证码输入框右侧的‘获取验证码’按钮")
                    time.sleep(1)
                    wait_for_popup = detect_popup_region()
                    if wait_for_popup:
                        close_popup(wait_for_popup)
                    new_code = None
                    for _ in range(5):
                        new_code = emailgetword.get_verification_code_once(check_limit=10)
                        if new_code and len(new_code) >= 4:
                            break
                        time.sleep(5)
                    GLOBAL_SMS_CODE_CACHE = new_code
                    print(f"🆕 短信码已更新为: {new_code}")
                    device_retry += 1
                    continue
                except Exception as e:
                    print(f"❌ main() 异常: {e}")
                    break

                if not success:
                    device_retry += 1
                    if device_retry < 2:
                        print(f"🔁 设备重试第 {device_retry} 次...")
                        time.sleep(3)

            if success:
                print(f"✅ {ne_name} 成功")
                _update_excel_status(EXCEL_PATH, excel_row, "是")
            else:
                print(f"❌ {ne_name} 失败")
                _update_excel_status(EXCEL_PATH, excel_row, "否")

            print("🔙 返回上一页...")
            pyautogui.press('esc')
            time.sleep(2)

        except Exception as e:
            print(f"❌ 处理 {ne_name} 出错: {e}")
            _update_excel_status(EXCEL_PATH, excel_row, f"异常")
            continue

    print("🎉 批量处理完成！")

# ========================
# 启动脚本
# ========================
if __name__ == "__main__":
    print("🚀 全自动密码修改脚本已启动")

    while True:
        print("\n" + "=" * 50)
        print("🔄 当前模式：支持多次连续执行")
        choice = input(
            "请选择操作：\n"
            "  🟢 输入 y → 执行一次密码修改\n"
            "  🔵 输入 e → 启动 Excel 批量处理\n"
            "  🔴 输入 q → 退出程序\n"
            "请输入: "
        ).strip().lower()

        if choice == 'y':
            print("🎯 开始执行单次密码修改...")
            print("👉 系统将在 3 秒后开始操作，请确保窗口已就绪！")
            time.sleep(3)
            try:
                result = main()
                if result == "y":
                    print("🎉 本次密码修改成功！")
                elif result == "x":
                    print("⚠️ 本次修改失败，可重试。")
                else:
                    print("❌ 本次执行异常或超时。")
            except Exception as e:
                print(f"❌ 执行 main() 时发生异常: {e}")
            input("↩️ 单次执行结束，按 Enter 键返回主菜单...")

        elif choice == 'e':
            print("📊 即将启动 Excel 批量处理...")
            confirm = input("👉 请确保 Excel 文件已关闭！确认开始？(y/n): ").strip().lower()
            if confirm == 'y':
                try:
                    time.sleep(5)
                    main_with_excel()
                except Exception as e:
                    print(f"❌ 批量处理出错: {e}")
            else:
                print("⏭️ 已取消批量处理。")
            input("↩️ 批量处理结束，按 Enter 键返回主菜单...")

        elif choice == 'q':
            print("👋 感谢使用，再见！")
            break
        else:
            print("❌ 输入无效，请输入 y、e 或 q。")
            time.sleep(1)
